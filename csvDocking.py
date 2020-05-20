import glob
import pandas as pd
import os
import openpyxl

data_path='C:\\workspace\\snowflake\\node\\output'

#処理するdirの決定
list=os.listdir(data_path)
print('select data source')
for i in range(0,len(list)):
    print(str(i)+':'+list[i])
process_dir=input()
process_dir=list[int(process_dir)]
print('start process:'+process_dir)
data_path=data_path+'\\'+process_dir
os.chdir(data_path)

#書き出しファイル
wbOut=openpyxl.Workbook()



# 同じフォルダのCSVファイルの一覧を取得
files = sorted(glob.glob('*.csv'))

# CSVファイルの数を取得
file_number = len(files)

# CSVファイルの中身をエクセル化する
csv_list = []

for file in files:
    data = pd.read_csv(file)
    refile=file.replace('.csv','')
    # Excel形式で出力
    data.to_excel(refile+'.xlsx', encoding='utf-8')

# エクセルファイルを集める
files = sorted(glob.glob('*.xlsx'))
rm_files = sorted(glob.glob('*.xlsx'))
 

for file in files:
    r=0
    c=1
    wbIn = openpyxl.load_workbook(file)
    sheetName=file.replace('.xlsx','')
    wsTo=wbOut.create_sheet(title=sheetName)
    wsFrom=wbIn.worksheets[0]
    for row in wsFrom:
        r=r+1
        c=1
        for cell in row:
            if(c!=1):
                wsTo.cell(row=r,column=c-1).value=cell.value
            c=c+1

# 結合完了のメッセージ
for file in rm_files:
    os.remove(file)
wbOut.save(data_path+'/merge.xlsx')


    
print(file_number,' 個のCSVファイルを結合して、merge.xlsxを作成しました。')
